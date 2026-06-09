import glob
import os
import re
import textwrap

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Always operate relative to this script's directory so it works regardless of CWD.
os.chdir(os.path.dirname(os.path.abspath(__file__)))

CLIENT_PREFIX = 'Leave the Key - '
INPUT_DIR = 'Input'
PARQUET_FILE = os.path.join(INPUT_DIR, 'leavethekey_numbers.parquet')
SOURCE_CSV = os.path.join(INPUT_DIR, 'leavethekey_numbers_2026-05-14.csv')

SOURCE_MAP = {
    # T1
    'T1Skiptrace': 'T1',
    'T1': 'T1',
    'Tier1': 'T1',
    'Tier 1': 'T1',
    'IDI': 'T1',
    'New source Launchskip': 'T1',
    'SkipGenie': 'T1',  # Leave the Key
    # T2
    'T2Skiptrace': 'T2',
    'T2': 'T2',
    'SkipForce': 'T2',
    'Skipforce': 'T2',  # casing variant (Leave the Key)
    # T3 (Batch/Skip)
    'Tier 3': 'T3',
    'Tier3': 'T3',
    'T3': 'T3',
    'BatchSkip': 'T3',  # Leave the Key
    # T6
    'T1.2': 'T6',
    'T6': 'T6',
    # T8
    'T8': 'T8',
    # Client-tracked source kept as its own category (Leave the Key)
    'ReiSift': 'ReiSift',
    'REISift': 'ReiSift',
    # Low-volume providers grouped
    'Locate Plus': 'Others',
    'Lead Sherpa': 'Others',
    'Zillow': 'Others',
    'CallRail': 'Others',
    'PropStream': 'Others',
    '1': 'Others',
}

# Dynamic: taken at script run time so age buckets always reflect "today".
TODAY = pd.Timestamp.today().normalize()
BUCKET_ORDER = ['<1 year', '1-2 years', '2+ years']

# Earliest plausible NUMBER CREATED AT per tier. Rows dated before a tier's floor
# cannot have come from that tier; if the source data claims they did (directly or
# via accumulated TAGS), we re-resolve them.
TIER_FLOOR = {
    'T3': pd.Timestamp('2025-04-01'),
    'T6': pd.Timestamp('2025-12-01'),
    'T8': pd.Timestamp('2026-03-01'),
}


def tier_allowed_on(tier, created_at):
    floor = TIER_FLOOR.get(tier)
    if floor is None or pd.isna(created_at):
        return True
    return created_at >= floor


def tier_launch_phrase(tiers):
    """E.g. ['T6', 'T8'] -> 'T6 in December 2025 and T8 in March 2026'."""
    parts = [f'{t} in {TIER_FLOOR[t].strftime("%B %Y")}' for t in tiers if t in TIER_FLOOR]
    if len(parts) <= 1:
        return parts[0] if parts else ''
    return ', '.join(parts[:-1]) + ' and ' + parts[-1]

# Cleaning_File_All_Numbers is CSV (Cava has 3M+ rows, exceeds the Excel sheet cap).
ALL_FILE = f'{CLIENT_PREFIX}Cleaning_File_All_Numbers.csv'
REPORT_INTERNAL_FILE = f'{CLIENT_PREFIX}Numbers Analysis Report - Internal.xlsx'
REPORT_CLIENT_FILE = f'{CLIENT_PREFIX}Numbers Analysis Report - Client.xlsx'
IMPORT_FILE = f'{CLIENT_PREFIX}Numbers Age Import File.xlsx'


def compile_or_load_parquet():
    """Load the combined dataset from the parquet cache.
    On first run, compile the source CSV into the parquet."""
    if os.path.exists(PARQUET_FILE):
        print(f'Loading from parquet: {PARQUET_FILE}')
        return pd.read_parquet(PARQUET_FILE)

    if not os.path.exists(SOURCE_CSV):
        raise FileNotFoundError(f'Source CSV not found: {SOURCE_CSV}')
    print(f'Compiling {SOURCE_CSV} into parquet (first run)...')
    with open(SOURCE_CSV, 'r', encoding='utf-8', errors='replace') as _f:
        _sep = ';' if ';' in _f.readline() else ','
    df = pd.read_csv(SOURCE_CSV, dtype=str, low_memory=False, sep=_sep)
    df.to_parquet(PARQUET_FILE, index=False)
    print(f'Saved {len(df):,} rows to {PARQUET_FILE}')
    return df


def dedup_output_name(n_rows):
    """{Prefix}Cleaning_File_Deduped_{N}k.xlsx with N rounded to nearest thousand."""
    n_k = round(n_rows / 1000)
    return f'{CLIENT_PREFIX}Cleaning_File_Deduped_{n_k}k.xlsx'


def source_from_tag_string(tags, created_at=None):
    """Recover a tier from the TAGS column when the SOURCE cell is blank.
    Order matters: T1.2 must win over T1 (substring), specific tiers before T1.
    TAGS in the CRM accumulates over time, so an old row can carry a newer tier's
    tag that wasn't its real origin. tier_allowed_on() blocks those mis-matches
    and lets the scan fall through to an older tier that fits the row's date."""
    if not isinstance(tags, str) or not tags.strip():
        return None
    s = tags.lower()
    if 't1.2' in s and tier_allowed_on('T6', created_at):
        return 'T6'
    if 't8' in s and tier_allowed_on('T8', created_at):
        return 'T8'
    if 't6' in s and tier_allowed_on('T6', created_at):
        return 'T6'
    if ('tier 3' in s or 't3' in s) and tier_allowed_on('T3', created_at):
        return 'T3'
    if 't2' in s:
        return 'T2'
    if 't1' in s or 'tier 1' in s or 'tier1' in s:
        return 'T1'
    return None


# Re-Skipped tag: when a property has been re-skiptraced, its TAGS column carries
# "Re-Skipped YYYY-MM". That month becomes the property's effective NUMBER CREATED AT,
# because the fresher contact data is what calling teams will actually use.
RESKIP_RE = re.compile(r'Re-?Skipped\s+(\d{4})-(\d{1,2})', re.IGNORECASE)


def extract_reskip_month(tags):
    if not isinstance(tags, str):
        return None
    m = RESKIP_RE.search(tags)
    if not m:
        return None
    return pd.Timestamp(year=int(m.group(1)), month=int(m.group(2)), day=1)


def apply_reskip_override(df):
    """For rows whose TAGS contain Re-Skipped YYYY-MM, overwrite NUMBER CREATED AT to
    YYYY-MM-01 and recompute Skipped Month-YYYY. Returns (new_df, mask_of_overridden)."""
    df = df.copy()
    reskip_dates = df['TAGS'].apply(extract_reskip_month)
    mask = reskip_dates.notna()
    if mask.any():
        df.loc[mask, 'NUMBER CREATED AT'] = pd.to_datetime(reskip_dates[mask])
        df.loc[mask, 'Skipped Month-YYYY'] = 'Skipped ' + reskip_dates[mask].dt.strftime('%Y-%m')
    return df, mask


def age_bucket(years):
    if pd.isna(years):
        return 'Unknown'
    if years < 1:
        return '<1 year'
    if years < 2:
        return '1-2 years'
    return '2+ years'


def clean(df):
    df['NUMBER CREATED AT'] = pd.to_datetime(df['NUMBER CREATED AT'])

    # Wrong-number flag is computed BEFORE STATUS fillna so the original 'Wrong Number' value drives detection.
    df['_is_wrong_number'] = df['STATUS'].astype(str).str.strip().str.lower() == 'wrong number'

    df['STATUS'] = df['STATUS'].fillna('Unknown')
    df.loc[df['STATUS'].astype(str).str.strip() == '', 'STATUS'] = 'Unknown'

    # Effective skip date for source attribution: a Re-Skipped YYYY-MM tag tells us
    # the property was re-skiptraced on that month, so the tier floor must use the
    # later of (raw NUMBER CREATED AT, Re-Skipped date) — otherwise a legit 2026 T6
    # re-skip on an old 2023 row would get invalidated just because the raw date
    # column hasn't been bumped yet (the bump happens later in apply_reskip_override).
    reskip_dates = df['TAGS'].apply(extract_reskip_month)
    effective_date = df['NUMBER CREATED AT'].copy()
    later = reskip_dates.notna() & (reskip_dates > effective_date.fillna(pd.Timestamp.min))
    effective_date.loc[later] = reskip_dates[later]

    mapped = df['SOURCE'].map(SOURCE_MAP)
    # Date-floor guard: a row dated before a tier's launch cannot be that tier,
    # even if the source CSV says so. Invalidate and let TAGS backfill re-resolve.
    for tier, floor in TIER_FLOOR.items():
        invalid = (mapped == tier) & effective_date.notna() & (effective_date < floor)
        if invalid.any():
            mapped.loc[invalid] = pd.NA
    needs_backfill = mapped.isna()
    if needs_backfill.any():
        mapped.loc[needs_backfill] = df.loc[needs_backfill].apply(
            lambda r: source_from_tag_string(r['TAGS'], effective_date.loc[r.name]), axis=1
        )
    df['SOURCE'] = mapped.fillna('Unknown')

    df['Skipped Month-YYYY'] = 'Skipped ' + df['NUMBER CREATED AT'].dt.strftime('%Y-%m')

    # Cava data has no ID column; FOLIO+ADDRESS+ZIP serves as the property identifier (per client).
    df['_property_key'] = (
        df['FOLIO'].fillna('').astype(str).str.strip() + '|' +
        df['ADDRESS'].fillna('').astype(str).str.strip() + '|' +
        df['ZIP'].fillna('').astype(str).str.strip()
    )
    return df


def build_breakdowns(properties):
    properties = properties.copy()
    properties['age_years'] = (TODAY - properties['NUMBER CREATED AT']).dt.days / 365.25
    properties['bucket'] = properties['age_years'].apply(age_bucket)

    age_counts = properties['bucket'].value_counts().reindex(BUCKET_ORDER).fillna(0).astype(int)

    sxa = pd.crosstab(properties['SOURCE'], properties['bucket'])
    sxa = sxa.reindex(columns=BUCKET_ORDER).fillna(0).astype(int)
    sxa['Total'] = sxa.sum(axis=1)
    sxa = sxa.sort_values('Total', ascending=False)

    status_counts = properties['STATUS'].value_counts()
    return age_counts, sxa, status_counts


# ---------- styling helpers ----------
TITLE_FONT = Font(name='Calibri', size=18, bold=True, color='FFFFFF')
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
SECTION_FONT = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
SECTION_FILL = PatternFill('solid', fgColor='2E75B6')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='5B9BD5')
TOTAL_FONT = Font(name='Calibri', size=11, bold=True)
TOTAL_FILL = PatternFill('solid', fgColor='DDEBF7')
NOTE_FONT = Font(name='Calibri', size=10, italic=True, color='595959')
INSIGHT_FONT = Font(name='Calibri', size=11, color='333333')
THIN = Side(style='thin', color='BFBFBF')
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)


def set_row(ws, row, values, font=None, fill=None, border=False, align=None, merge_to=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        if font: c.font = font
        if fill: c.fill = fill
        if border: c.border = CELL_BORDER
        if align: c.alignment = align
    if merge_to:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merge_to)


def write_title(ws, row, text, cols):
    set_row(ws, row, [text], font=TITLE_FONT, fill=TITLE_FILL, align=CENTER, merge_to=cols)
    ws.row_dimensions[row].height = 28


def write_section(ws, row, text, cols):
    set_row(ws, row, [text], font=SECTION_FONT, fill=SECTION_FILL, align=Alignment(horizontal='left', vertical='center'), merge_to=cols)
    ws.row_dimensions[row].height = 22


def estimate_row_height(text, merged_col_width, line_height=16, padding=8):
    """Estimate row height needed to fit wrapped text inside merged cells."""
    if not text:
        return 18
    # Conservative: 0.9 × column width units ≈ characters per line at Calibri 11pt,
    # accounting for wide chars like em-dashes and digits.
    chars_per_line = max(20, int(merged_col_width * 0.9))
    total_lines = 0
    for segment in str(text).split('\n'):
        if not segment:
            total_lines += 1
            continue
        wrapped = textwrap.wrap(segment, width=chars_per_line, break_long_words=False, break_on_hyphens=False)
        total_lines += max(1, len(wrapped))
    return max(20, total_lines * line_height + padding)


def write_note(ws, row, text, cols, merged_col_width):
    set_row(ws, row, [text], font=NOTE_FONT, align=LEFT_WRAP, merge_to=cols)
    ws.row_dimensions[row].height = estimate_row_height(text, merged_col_width, line_height=15, padding=6)


def write_insight(ws, row, text, cols, merged_col_width):
    set_row(ws, row, [text], font=INSIGHT_FONT, align=LEFT_WRAP, merge_to=cols)
    ws.row_dimensions[row].height = estimate_row_height(text, merged_col_width, line_height=16, padding=8)


def write_age_table(ws, start_row, age_counts, label):
    total = int(age_counts.sum())
    set_row(ws, start_row, ['Data age', f'{label}', '% of total'],
            font=HEADER_FONT, fill=HEADER_FILL, border=True, align=CENTER)
    r = start_row + 1
    for bucket in BUCKET_ORDER:
        cnt = int(age_counts.get(bucket, 0))
        pct = (cnt / total * 100) if total else 0
        set_row(ws, r, [bucket, cnt, f'{pct:.1f}%'], border=True, align=CENTER)
        r += 1
    set_row(ws, r, ['Total', total, '100.0%'], font=TOTAL_FONT, fill=TOTAL_FILL, border=True, align=CENTER)
    return r + 1


def write_sxa_table(ws, start_row, sxa):
    cols = ['SOURCE'] + BUCKET_ORDER + ['Total']
    set_row(ws, start_row, cols, font=HEADER_FONT, fill=HEADER_FILL, border=True, align=CENTER)
    r = start_row + 1
    for src, row in sxa.iterrows():
        values = [src] + [int(row.get(b, 0)) for b in BUCKET_ORDER] + [int(row['Total'])]
        set_row(ws, r, values, border=True, align=CENTER)
        r += 1
    totals = ['Total'] + [int(sxa[b].sum()) for b in BUCKET_ORDER] + [int(sxa['Total'].sum())]
    set_row(ws, r, totals, font=TOTAL_FONT, fill=TOTAL_FILL, border=True, align=CENTER)
    return r + 1


def write_status_table(ws, start_row, status_counts):
    total = int(status_counts.sum())
    set_row(ws, start_row, ['STATUS', 'Properties', '% of total'],
            font=HEADER_FONT, fill=HEADER_FILL, border=True, align=CENTER)
    r = start_row + 1
    for status, cnt in status_counts.items():
        pct = (cnt / total * 100) if total else 0
        set_row(ws, r, [str(status), int(cnt), f'{pct:.1f}%'], border=True, align=CENTER)
        r += 1
    set_row(ws, r, ['Total', total, '100.0%'], font=TOTAL_FONT, fill=TOTAL_FILL, border=True, align=CENTER)
    return r + 1


def set_column_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- main ----------
df = compile_or_load_parquet()
df = clean(df)

# Columns we add for internal bookkeeping; never written to deliverables.
INTERNAL_COLS = ['_is_wrong_number', '_property_key']

df.drop(columns=INTERNAL_COLS).to_csv(ALL_FILE, index=False)
print(f'All Numbers: {len(df):,} rows saved to {ALL_FILE}')

# Dedup: one row per property (FOLIO+ADDRESS+ZIP) keeping OLDEST NUMBER CREATED AT
df_sorted = df.sort_values('NUMBER CREATED AT', ascending=True, kind='mergesort')
df_dedup = df_sorted.drop_duplicates(subset=['_property_key'], keep='first').sort_index()

# Property-level Re-Skipped override: if the deduped row carries a Re-Skipped YYYY-MM tag,
# treat that month as the property's effective skip date. Drives reports + import file.
df_dedup, reskip_mask = apply_reskip_override(df_dedup)
client_has_reskip = bool(reskip_mask.any())
if client_has_reskip:
    print(f'Re-Skipped tag detected on {int(reskip_mask.sum()):,} properties — '
          f'dates overridden, import file restricted to age < 1 year.')

DEDUP_FILE = dedup_output_name(len(df_dedup))
df_dedup.drop(columns=INTERNAL_COLS).to_excel(DEDUP_FILE, index=False)
print(f'Deduped (oldest per property): {len(df_dedup):,} rows saved to {DEDUP_FILE}')

# Import file: folio, address, zip + tag (the Skipped YYYY-MM value from the deduped row).
# When the client has been through this process before (Re-Skipped tags present), exclude
# anything 1+ year old — those properties were already tagged in 8020rei domain on the
# prior iteration and don't need to be re-imported.
df_import_src = df_dedup
if client_has_reskip:
    age_years_import = (TODAY - df_dedup['NUMBER CREATED AT']).dt.days / 365.25
    df_import_src = df_dedup[age_years_import < 1]
df_import = df_import_src[['FOLIO', 'ADDRESS', 'ZIP', 'Skipped Month-YYYY']].rename(
    columns={'FOLIO': 'folio', 'ADDRESS': 'address', 'ZIP': 'zip', 'Skipped Month-YYYY': 'tag'}
)
df_import.to_excel(IMPORT_FILE, index=False)
print(f'Import file: {len(df_import):,} rows saved to {IMPORT_FILE}')

# ---------- breakdowns ----------
# Cava: wrong-number properties are identified by STATUS == 'Wrong Number' (TAGS-based detection used for Kentucky doesn't apply).
wn_rows = df[df['_is_wrong_number']].copy()
wn_sorted = wn_rows.sort_values('NUMBER CREATED AT', ascending=True, kind='mergesort')
wn_dedup = wn_sorted.drop_duplicates(subset=['_property_key'], keep='first')
wn_dedup, _ = apply_reskip_override(wn_dedup)
wn_age, wn_sxa, _ = build_breakdowns(wn_dedup)

# Full portfolio view = the entire deduped dataset (one row per property).
# Wrong-number properties are INCLUDED here so the totals match the deduped file;
# they are also detailed separately on the Wrong Numbers sheet.
df_full = df_dedup.copy()
full_age, full_sxa, _ = build_breakdowns(df_full)

total_properties = len(df_full)
wn_properties = len(wn_dedup)
wn_rate = wn_properties / total_properties * 100
over1_full = int(full_age.get('1-2 years', 0) + full_age.get('2+ years', 0))
over2_full = int(full_age.get('2+ years', 0))
over1_wn = int(wn_age.get('1-2 years', 0) + wn_age.get('2+ years', 0))
over2_wn = int(wn_age.get('2+ years', 0))

date_min = df_dedup['NUMBER CREATED AT'].min().strftime('%Y-%m-%d')
date_max = df_dedup['NUMBER CREATED AT'].max().strftime('%Y-%m-%d')

# ---------- workbook ----------
COLS = 6
COL_WIDTHS = [30, 18, 14, 14, 14, 14]
MERGED_WIDTH = sum(COL_WIDTHS[:COLS])


def build_full_portfolio_sheet(ws, include_sources):
    set_column_widths(ws, COL_WIDTHS)
    r = 1
    write_title(ws, r, 'Skiptrace Numbers — Full Portfolio Analysis', COLS); r += 1
    write_note(ws, r, f'Generated {TODAY.strftime("%Y-%m-%d")} · One record per property · Includes all properties (wrong-number properties also detailed on the separate sheet) · Data range: {date_min} to {date_max}', COLS, MERGED_WIDTH); r += 2

    write_section(ws, r, 'Summary', COLS); r += 1
    set_row(ws, r, ['Total properties in portfolio', total_properties], border=True); r += 1
    set_row(ws, r, ['Properties with data 1+ years old', f'{over1_full:,}  ({over1_full/total_properties*100:.1f}%)'], border=True); r += 1
    set_row(ws, r, ['Properties with data 2+ years old', f'{over2_full:,}  ({over2_full/total_properties*100:.1f}%)'], border=True); r += 1
    set_row(ws, r, ['Wrong-number properties (also on separate sheet)', f'{wn_properties:,}  ({wn_rate:.1f}% of portfolio)'], border=True); r += 2

    write_section(ws, r, 'Data age distribution', COLS); r += 1
    write_note(ws, r, 'Each property counted once. Older contact data typically yields lower connect rates because phone numbers change over time.', COLS, MERGED_WIDTH); r += 1
    r = write_age_table(ws, r, full_age, 'Properties')
    r += 1

    if include_sources:
        write_section(ws, r, 'Source × Age matrix', COLS); r += 1
        write_note(ws, r, 'Skiptrace tier each property came from, cross-referenced with data age. Use to prioritize which source batches to refresh.', COLS, MERGED_WIDTH); r += 1
        r = write_sxa_table(ws, r, full_sxa)
        r += 1

    write_section(ws, r, 'Insights & recommendations', COLS); r += 1
    if include_sources:
        t1_count = int(full_sxa.loc['T1', 'Total']) if 'T1' in full_sxa.index else 0
        t1_share = t1_count / total_properties * 100 if total_properties else 0
        t1_2plus = int(full_sxa.loc['T1', '2+ years']) if 'T1' in full_sxa.index else 0
        t6_count = int(full_sxa.loc['T6', 'Total']) if 'T6' in full_sxa.index else 0
        t8_count = int(full_sxa.loc['T8', 'Total']) if 'T8' in full_sxa.index else 0
        insights = [
            f'1. Portfolio of {total_properties:,} properties. {over1_full:,} ({over1_full/total_properties*100:.1f}%) have data 1+ years old; {over2_full:,} ({over2_full/total_properties*100:.1f}%) are 2+ years old. Older data tends to lower connect rates.',
            f'2. T1 dominates at {t1_count:,} properties ({t1_share:.1f}%). This reflects sourcing history: T1 and T2 are the long-standing tiers, while newer tiers launched later ({tier_launch_phrase(["T3", "T6", "T8"])}), so T1 carries the oldest data.',
            f'3. The {t1_2plus:,} T1 records that are 2+ years old are the most likely driver of softer contact results on this cohort — phone numbers gathered that long ago have had the most time to go stale.',
            f'4. T6 ({t6_count:,}, launched {TIER_FLOOR["T6"].strftime("%B %Y")}) and T8 ({t8_count:,}, launched {TIER_FLOOR["T8"].strftime("%B %Y")}) show the expected fresh-data profile. Treat them as the higher-intent cohort in active calling queues.',
            '5. Recommended action: prioritize a re-skiptrace of T1 records that are 2+ years old. Largest cohort and most likely to yield recovered contacts from a fresh pull.',
        ]
    else:
        under1 = int(full_age.get('<1 year', 0))
        insights = [
            f'1. Portfolio of {total_properties:,} properties. {over1_full:,} ({over1_full/total_properties*100:.1f}%) have data 1+ years old; {over2_full:,} ({over2_full/total_properties*100:.1f}%) are 2+ years old. Older contact data typically produces lower connect rates as phone numbers change over time.',
            f'2. The 2+ year cohort is the largest single group at {over2_full:,} properties ({over2_full/total_properties*100:.1f}%). These records have had the most time to go stale and are the most likely to produce disconnected or wrong numbers on call.',
            f'3. Roughly {under1:,} properties ({under1/total_properties*100:.1f}%) have data under 1 year old. Treat these as the highest-intent cohort and prioritize them in active calling queues.',
            '4. Recommended action: schedule a periodic refresh of records 2+ years old. This is the cohort most likely to yield recovered contacts from a fresh data pull.',
            '5. Going forward, maintaining a rolling refresh cadence will keep the age profile weighted toward fresher, higher-yield records over time.',
        ]
    for text in insights:
        write_insight(ws, r, text, COLS, MERGED_WIDTH); r += 1

    ws.freeze_panes = 'A4'


def build_wrong_numbers_sheet(ws, include_sources):
    set_column_widths(ws, COL_WIDTHS)
    r = 1
    write_title(ws, r, 'Skiptrace Numbers — Wrong-Number Analysis', COLS); r += 1
    write_note(ws, r, 'Properties with at least one phone number flagged as a wrong number · One record per property', COLS, MERGED_WIDTH); r += 2

    if wn_properties == 0:
        write_section(ws, r, 'Summary', COLS); r += 1
        set_row(ws, r, ['Wrong-number properties', 0], border=True); r += 2
        write_insight(ws, r,
            f'No properties in this portfolio carry a wrong-number flag (STATUS = "Wrong Number"). '
            f'Reviewed {total_properties:,} deduped properties.',
            COLS, MERGED_WIDTH)
        ws.freeze_panes = 'A4'
        return

    pct_over1 = over1_wn / wn_properties * 100
    pct_over2 = over2_wn / wn_properties * 100

    write_section(ws, r, 'Summary', COLS); r += 1
    set_row(ws, r, ['Wrong-number properties', wn_properties], border=True); r += 1
    set_row(ws, r, ['Share of portfolio', f'{wn_rate:.2f}%  ({wn_properties:,} of {total_properties:,})'], border=True); r += 1
    set_row(ws, r, ['With data 1+ years old', f'{over1_wn:,}  ({pct_over1:.1f}%)'], border=True); r += 1
    set_row(ws, r, ['With data 2+ years old', f'{over2_wn:,}  ({pct_over2:.1f}%)'], border=True); r += 2

    write_section(ws, r, 'Data age distribution (wrong-number properties)', COLS); r += 1
    write_note(ws, r, 'Older data is more likely to reflect stale numbers that a fresh skiptrace can recover.', COLS, MERGED_WIDTH); r += 1
    r = write_age_table(ws, r, wn_age, 'Properties')
    r += 1

    if include_sources:
        write_section(ws, r, 'Source × Age matrix (wrong numbers only)', COLS); r += 1
        write_note(ws, r, 'Source tier of the wrong numbers, split by data age. Tiers with heavy older-data footprint are the strongest re-skiptrace candidates.', COLS, MERGED_WIDTH); r += 1
        r = write_sxa_table(ws, r, wn_sxa)
        r += 1

    write_section(ws, r, 'Insights & recommendations', COLS); r += 1
    wn_recent = int(wn_age.get('<1 year', 0))
    if include_sources:
        wn_t1 = int(wn_sxa.loc['T1', 'Total']) if 'T1' in wn_sxa.index else 0
        wn_t1_share = wn_t1 / wn_properties * 100 if wn_properties else 0
        insights = [
            f'1. Volume: {wn_properties:,} wrong-number properties ({wn_rate:.2f}% of the original portfolio). Manageable scale, but worth addressing to protect caller productivity.',
            f'2. Age profile: {over1_wn:,} of {wn_properties:,} ({over1_wn/wn_properties*100:.1f}%) have data 1+ years old. Stale data is the likely root cause; a fresh skiptrace should recover usable numbers for most.',
            f'3. T1 drives wrong numbers ({wn_t1:,} of {wn_properties:,}, {wn_t1_share:.1f}%). Consistent with T1 being the longest-standing tier — newer tiers launched later ({tier_launch_phrase(["T3", "T6", "T8"])}) — so T1 data has had the most time to go stale.',
            f'4. Recent wrong numbers (<1 year): {wn_recent:,}. Less likely to be fixed by re-skiptracing; worth reviewing at the originating source for data quality.',
            '5. Suggested workflow: (a) exclude wrong-number properties from active calling queues, (b) submit the 1+ year cohort to a fresh skiptrace, (c) re-ingest returned numbers and recycle into the dialer.',
        ]
    else:
        insights = [
            f'1. Volume: {wn_properties:,} wrong-number properties ({wn_rate:.2f}% of the original portfolio). Manageable scale, but worth addressing to protect caller productivity.',
            f'2. Age profile: {over1_wn:,} of {wn_properties:,} ({over1_wn/wn_properties*100:.1f}%) have data 1+ years old. Stale data is the most likely root cause — a fresh skiptrace should recover usable numbers for a large share of these.',
            f'3. The 2+ year cohort is the biggest slice of wrong numbers at {over2_wn:,} properties ({over2_wn/wn_properties*100:.1f}%). This is where a refresh is most likely to pay off.',
            f'4. Recent wrong numbers (<1 year): {wn_recent:,} properties. These are less likely to be resolved by re-skiptracing and are worth reviewing for underlying data quality.',
            '5. Suggested workflow: (a) exclude wrong-number properties from active calling queues, (b) submit the 1+ year cohort to a fresh skiptrace, (c) re-ingest returned numbers and recycle into the dialer.',
        ]
    for text in insights:
        write_insight(ws, r, text, COLS, MERGED_WIDTH); r += 1

    ws.freeze_panes = 'A4'


def build_report(path, include_sources):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Full Portfolio'
    build_full_portfolio_sheet(ws1, include_sources)
    ws2 = wb.create_sheet('Wrong Numbers')
    build_wrong_numbers_sheet(ws2, include_sources)
    wb.save(path)
    print(f'Report saved to {path}')


build_report(REPORT_INTERNAL_FILE, include_sources=True)
build_report(REPORT_CLIENT_FILE, include_sources=False)
